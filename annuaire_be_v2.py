#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
annuaire_be_v4.py - Base de prospection NATIONALE des bureaux d'etudes energie
Source principale : OPEN DATA RGE ADEME (API DataFair, licence ouverte, France entiere)
Sortie : base_be_tertiaire.xlsx (style, multi-onglets) + CSV + run_summary.json

Pourquoi l'ADEME et plus le scraping OPQIBI :
  - couverture nationale native (fini le bug region)
  - email + telephone + site deja dans la donnee
  - champ meta_domaine structure (fini le bug d'extraction des qualifs)
  - ~1 min via API au lieu de 2h de scan
  - Licence Ouverte Etalab : revente commerciale autorisee

Pipeline :
  1. PULL ADEME : toutes les lignes meta_domaine="Etudes energetiques"
     (exclusion Architecte + domaines travaux/installations), dedup SIRET.
  2. ENRICHISSEMENT gratuit (sans cle) :
     - recherche-entreprises.api.gouv.fr : etat admin + effectif + dirigeant
     - BODACC opendatasoft : procedures collectives
     - sites web : email pour les fiches sans email
  3. SCORING + tiers (CHAUD / TIEDE / A QUALIFIER / HORS CIBLE / EN DIFFICULTE / FERMEE)
     avec bonus "domaine tertiaire" (logement collectif, enveloppe, systeme technique...).
  4. EXPORTS.

Usage :
  pip install requests beautifulsoup4 pandas openpyxl
  python annuaire_be_v4.py                     # national complet
  python annuaire_be_v4.py --dep 69            # un departement
  python annuaire_be_v4.py --skip-enrich       # sans SIRENE/BODACC/sites (rapide)
  python annuaire_be_v4.py --from-csv base.csv # reprendre un CSV, sauter le pull
"""

import argparse
import csv
import json
import re
import time
import unicodedata
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path

import pandas as pd
import requests

HEADERS = {"User-Agent": "Mozilla/5.0 (compatible; base-be-tertiaire/4.0; usage pro)"}
DEPS_AURA = {"01", "03", "07", "15", "26", "38", "42", "43", "63", "69", "73", "74"}

ADEME_LINES = "https://data.ademe.fr/data-fair/api/v1/datasets/liste-des-entreprises-rge-2/lines"

TIER_HOT, TIER_WARM, TIER_QUAL = "CHAUD", "TIEDE", "A QUALIFIER"
TIER_OUT, TIER_DIFF, TIER_DEAD = "HORS CIBLE", "EN DIFFICULTE", "FERMEE"
TIER_ORDER = [TIER_HOT, TIER_WARM, TIER_QUAL, TIER_OUT, TIER_DIFF, TIER_DEAD]

# domaines d'etudes les plus pertinents pour le decret tertiaire (bonus scoring)
TERTIAIRE_KW = ["logement collectif", "tertiaire", "enveloppe", "systeme technique",
                "thermique reglementaire", "eclairage", "acv", "commiss", "photovolt"]


def norm(s):
    s = unicodedata.normalize("NFD", str(s or "")).encode("ascii", "ignore").decode()
    return s.lower().strip()


def domaine_tertiaire(dom):
    n = norm(dom)
    return any(k in n for k in TERTIAIRE_KW)

# ================================================================ PULL ADEME

def pull_ademe(session, dep=None):
    """Recupere les lignes RGE via l'API DataFair, pagination par curseur 'next'.

    Pas de filtre serveur (source de 0 resultat quand la syntaxe qs ne passe pas) :
    on rapatrie et on filtre 100% cote client (meta_domaine + departement), fiable.
    """
    print("== Pull open data RGE ADEME ==", flush=True)
    params = {"size": 5000}
    rows, url, err_streak, page = [], ADEME_LINES, 0, 0
    first = True
    while url:
        try:
            r = session.get(url, params=params if first else None,
                            headers=HEADERS, timeout=90)
            r.raise_for_status()
        except requests.RequestException as e:
            print(f"  ADEME erreur : {e}", flush=True)
            err_streak += 1
            if err_streak >= 3:
                print("  3 erreurs consecutives, arret.", flush=True)
                break
            time.sleep(3)
            continue
        err_streak = 0
        first = False

        data = r.json()
        results = data.get("results", [])
        rows.extend(results)
        page += 1
        total = data.get("total")
        print(f"  page {page} : +{len(results)} (total {len(rows)}"
              + (f" / {total}" if total else "") + ")", flush=True)
        url = data.get("next")
        if not results or page >= 400:
            break
        time.sleep(0.15)

    df = pd.DataFrame(rows)
    if df.empty:
        print("!! ADEME : 0 ligne recuperee (verifier l'acces reseau a data.ademe.fr)")
        return df
    print(f"  brut : {len(df)} labels, colonnes : {list(df.columns)[:12]}...")

    # normaliser les noms de colonnes attendus (tolerant aux variations)
    ren = {}
    for c in df.columns:
        cl = c.lower()
        if cl == "nom_entreprise": ren[c] = "nom"
        elif cl == "commune": ren[c] = "ville"
        elif cl == "site_internet": ren[c] = "site"
    df = df.rename(columns=ren)
    for need in ("meta_domaine", "domaine", "email", "telephone", "site",
                 "siret", "nom", "adresse", "code_postal", "ville", "organisme"):
        if need not in df.columns:
            df[need] = ""

    # filtre etudes energetiques + exclusion architecte
    md = df["meta_domaine"].map(norm)
    dom = df["domaine"].map(norm)
    df = df[(md == "etudes energetiques") & (dom != "architecte")].copy()
    print(f"  filtre 'Etudes energetiques' (excl. archi) : {len(df)} labels")

    # regrouper les domaines par entreprise (une entreprise = plusieurs labels)
    df["siret"] = df["siret"].astype(str).str.replace(r"\D", "", regex=True).str[:14]
    df["siren"] = df["siret"].str[:9]
    agg = {c: "first" for c in ["nom", "email", "telephone", "site", "adresse",
                                "code_postal", "ville", "siren", "organisme"]}
    domaines = df.groupby("siret")["domaine"].apply(lambda s: ";".join(sorted(set(s.dropna()))))
    base = df.groupby("siret").agg(agg)
    base["domaines_rge"] = domaines
    base = base.reset_index()
    base = base[base["siret"].str.len() == 14]

    if dep:
        base = base[base["code_postal"].astype(str).str[:2] == str(dep)]
        print(f"  filtre departement {dep} : {len(base)}")

    print(f"  entreprises uniques (dedup SIRET) : {len(base)}")
    return base

# ================================================================ ENRICHISSEMENT

def _sirene_one(args):
    session, siren = args
    if len(str(siren)) != 9:
        return "", "", ""
    try:
        r = session.get("https://recherche-entreprises.api.gouv.fr/search",
                        params={"q": siren, "per_page": 1}, headers=HEADERS, timeout=15)
        if r.status_code == 200:
            res = r.json().get("results", [])
            if res:
                e = res[0]
                etat = e.get("etat_administratif", "") or ""
                eff = e.get("tranche_effectif_salarie", "") or ""
                diri = ""
                dg = e.get("dirigeants", [])
                if dg:
                    d0 = dg[0]
                    diri = " ".join(x for x in [d0.get("prenoms", ""),
                             d0.get("nom", d0.get("denomination", ""))] if x).strip()
                return etat, eff, diri
    except requests.RequestException:
        pass
    return "", "", ""


def enrich_sirene(session, df):
    print("== Enrichissement SIRENE (parallele) ==", flush=True)
    sirens = list(df["siren"].fillna(""))
    results = [None] * len(sirens)
    with ThreadPoolExecutor(max_workers=10) as ex:
        for i, res in enumerate(ex.map(_sirene_one, [(session, s) for s in sirens])):
            results[i] = res
            if (i + 1) % 500 == 0:
                print(f"  {i + 1}/{len(sirens)}", flush=True)
    df["etat_sirene"] = [r[0] for r in results]
    df["tranche_effectif"] = [r[1] for r in results]
    df["dirigeant"] = [r[2] for r in results]
    return df


def _bodacc_one(args):
    session, siren = args
    if len(str(siren)) != 9:
        return ""
    url = "https://bodacc-datadila.opendatasoft.com/api/records/1.0/search/"
    try:
        r = session.get(url, params={"dataset": "annonces-commerciales",
                                     "q": siren, "rows": 5, "sort": "dateparution"},
                        headers=HEADERS, timeout=15)
        if r.status_code == 200:
            for rec in r.json().get("records", []):
                f = rec.get("fields", {})
                if "collective" in (f.get("familleavis_lib") or "").lower():
                    return f"{f.get('familleavis_lib')} ({str(f.get('dateparution',''))[:10]})"
    except requests.RequestException:
        pass
    return ""


def enrich_bodacc(session, df):
    print("== Enrichissement BODACC (parallele) ==", flush=True)
    sirens = list(df["siren"].fillna(""))
    procs = [None] * len(sirens)
    with ThreadPoolExecutor(max_workers=8) as ex:
        for i, res in enumerate(ex.map(_bodacc_one, [(session, s) for s in sirens])):
            procs[i] = res
            if (i + 1) % 500 == 0:
                print(f"  {i + 1}/{len(sirens)}", flush=True)
    df["procedure_collective"] = procs
    return df


EMAIL_RE = re.compile(r"[\w.\-]+@[\w.\-]+\.[a-z]{2,}", re.I)
BAD_MAIL = ("example", "wixpress", "sentry", "@2x", ".png", ".jpg", "@sentry")


def _email_one(args):
    session, idx, site = args
    site = str(site).strip()
    if not site.startswith("http"):
        site = "https://" + site
    for path in ("", "/contact", "/mentions-legales"):
        try:
            r = session.get(site.rstrip("/") + path, headers=HEADERS,
                            timeout=8, allow_redirects=True)
            if r.status_code != 200 or not r.headers.get("content-type", "").startswith(("text", "application/xhtml")):
                continue
            for mail in EMAIL_RE.findall(r.text[:500000]):
                if not any(b in mail.lower() for b in BAD_MAIL):
                    return idx, mail
        except Exception:
            continue
    return idx, ""


def enrich_emails_sites(session, df):
    print("== Enrichissement emails via sites web (parallele) ==", flush=True)
    if "email_source" not in df.columns:
        df["email_source"] = ""
    todo = df[(~df["email"].astype(str).str.contains("@")) & (df["site"].astype(str).str.len() > 3)]
    print(f"  {len(todo)} fiches sans email mais avec site", flush=True)
    tasks = [(session, idx, row["site"]) for idx, row in todo.iterrows()]
    done = 0
    with ThreadPoolExecutor(max_workers=12) as ex:
        for idx, mail in ex.map(_email_one, tasks):
            if mail:
                df.at[idx, "email"] = mail
                df.at[idx, "email_source"] = "site_web"
            done += 1
            if done % 200 == 0:
                print(f"  {done}/{len(todo)}", flush=True)
    return df

# ================================================================ SCORING

EFF_MAP = {  # codes tranche effectif INSEE -> borne haute approx
    "NN": 0, "00": 0, "01": 2, "02": 5, "03": 9, "11": 19, "12": 49,
    "21": 99, "22": 199, "31": 249, "32": 499, "41": 999, "42": 1999,
    "51": 4999, "52": 9999, "53": 10000,
}


def score_row(r):
    pts, why = 3, ["RGE etudes"]  # deja filtre etudes energetiques
    if "@" in str(r.get("email", "")):
        pts += 2; why.append("email")
    if len(str(r.get("telephone", ""))) > 5:
        pts += 1; why.append("tel")
    if len(str(r.get("site", ""))) > 3:
        pts += 1; why.append("site")
    if domaine_tertiaire(r.get("domaines_rge", "")):
        pts += 2; why.append("domaine tertiaire")
    eff = EFF_MAP.get(str(r.get("tranche_effectif", "")), 0)
    if 1 <= eff <= 49:
        pts += 2; why.append("effectif cible")
    return pts, ", ".join(why), eff


def tier_row(r):
    if r.get("etat_sirene") == "C":
        return TIER_DEAD
    if str(r.get("procedure_collective", "")).strip():
        return TIER_DIFF
    if r["effectif_num"] > 250:
        return TIER_OUT
    if r["score"] >= 8:
        return TIER_HOT
    if r["score"] >= 6:
        return TIER_WARM
    return TIER_QUAL

# ================================================================ EXCEL

_ILLEGAL = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")  # caracteres de controle interdits par Excel

def _clean(v):
    return _ILLEGAL.sub("", str(v if v is not None else ""))

def build_excel(df, path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    HDR_FILL = PatternFill("solid", start_color="0E8E73")
    HDR_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    BODY = Font(name="Arial", size=10)
    TIER_FILL = {
        TIER_HOT: PatternFill("solid", start_color="FFE0CC"),
        TIER_WARM: PatternFill("solid", start_color="FFF3D6"),
        TIER_QUAL: PatternFill("solid", start_color="E8F1FA"),
        TIER_OUT: PatternFill("solid", start_color="F2F2F2"),
        TIER_DIFF: PatternFill("solid", start_color="FADBD8"),
        TIER_DEAD: PatternFill("solid", start_color="E6B8B4"),
    }
    thin = Border(bottom=Side(style="thin", color="E3E9EF"))

    cols = [("tier", "Tier", 14), ("score", "Score", 7), ("nom", "Nom", 34),
            ("email", "Email", 30), ("telephone", "Telephone", 14), ("site", "Site", 26),
            ("dirigeant", "Dirigeant", 26), ("ville", "Ville", 18), ("departement", "Dep", 6),
            ("domaines_rge", "Domaines RGE etudes", 40), ("tranche_effectif", "Eff.", 7),
            ("procedure_collective", "Procedure collective", 22), ("etat_sirene", "SIRENE", 8),
            ("siret", "SIRET", 16), ("adresse", "Adresse", 30), ("code_postal", "CP", 7),
            ("organisme", "Organisme", 12), ("email_source", "Src email", 10),
            ("score_detail", "Detail score", 40)]

    wb = Workbook()

    def sheet(ws, data, title):
        ws.title = title
        for j, (_, label, w) in enumerate(cols, 1):
            c = ws.cell(1, j, label); c.font, c.fill = HDR_FONT, HDR_FILL
            c.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(j)].width = w
        for i, (_, r) in enumerate(data.iterrows(), 2):
            f = TIER_FILL.get(r["tier"])
            for j, (k, _, _) in enumerate(cols, 1):
                c = ws.cell(i, j, _clean(r.get(k, ""))); c.font, c.border = BODY, thin
                if f: c.fill = f
        ws.freeze_panes = "C2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{max(2, len(data)+1)}"

    live = df[~df["tier"].isin([TIER_DIFF, TIER_DEAD])]
    dead = df[df["tier"].isin([TIER_DIFF, TIER_DEAD])]
    sheet(wb.active, live, "Prospection")
    sheet(wb.create_sheet(), dead, "Ecartes")

    st = wb.create_sheet("Stats")
    st.column_dimensions["A"].width = 38; st.column_dimensions["B"].width = 14
    st["A1"], st["B1"] = "Indicateur", "Valeur"
    for c in ("A1", "B1"):
        st[c].font, st[c].fill = HDR_FONT, HDR_FILL
    n = len(live) + 1
    for i, (k, v) in enumerate([
        ("Bureaux d'etudes (prospection)", f"=COUNTA(Prospection!C2:C{n})"),
        ("Avec email", f'=COUNTIF(Prospection!D2:D{n},"*@*")'),
        ("Avec telephone", f'=COUNTIF(Prospection!E2:E{n},"?*")'),
        ("CHAUD", f'=COUNTIF(Prospection!A2:A{n},"{TIER_HOT}")'),
        ("TIEDE", f'=COUNTIF(Prospection!A2:A{n},"{TIER_WARM}")'),
        ("Ecartes (fermees + difficultes)", f"=COUNTA(Ecartes!C2:C{max(2,len(dead)+1)})"),
    ], 2):
        st[f"A{i}"], st[f"B{i}"] = k, v
        st[f"A{i}"].font = st[f"B{i}"].font = BODY
    wb.calculation.fullCalcOnLoad = True
    wb.save(path)

# ================================================================ MAIN

def _need_enrich(df):
    """Marque les lignes pas encore enrichies (colonne 'enrichi' != '1')."""
    if "enrichi" not in df.columns:
        df["enrichi"] = ""
    return df["enrichi"].astype(str) != "1"


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dep", help="limiter a un departement (ex: 69)")
    ap.add_argument("--from-csv", help="reprendre un CSV existant (sauter le pull ADEME)")
    ap.add_argument("--skip-enrich", action="store_true")
    ap.add_argument("--lot", type=int, default=1500,
                    help="nombre de BE a enrichir par execution (reprise sur plusieurs runs)")
    ap.add_argument("--max-minutes", type=float, default=100,
                    help="arret propre avant cette duree (timeout GitHub = 120 min)")
    args = ap.parse_args()

    t0 = time.time()
    session = requests.Session()
    out = Path(".")
    cache = out / "base_be_max_national.csv"

    # 1) Charger : reprise du cache si present, sinon pull ADEME
    if args.from_csv and Path(args.from_csv).exists():
        df = pd.read_csv(args.from_csv, dtype=str).fillna("")
        print(f"CSV repris (--from-csv) : {len(df)} lignes", flush=True)
    elif cache.exists() and not args.dep:
        df = pd.read_csv(cache, dtype=str).fillna("")
        print(f"Cache repris ({cache.name}) : {len(df)} lignes -> on continue l'enrichissement", flush=True)
    else:
        df = pull_ademe(session, dep=args.dep)
        if df.empty:
            return

    for col in ("email", "site", "siren", "email_source", "etat_sirene",
                "procedure_collective", "dirigeant", "tranche_effectif",
                "domaines_rge", "enrichi"):
        if col not in df.columns:
            df[col] = ""
    df["departement"] = df["code_postal"].astype(str).str[:2]

    # sauvegarde immediate du socle (avant tout enrichissement) pour ne jamais repartir de zero
    df.to_csv(cache, index=False, quoting=csv.QUOTE_MINIMAL)

    if not args.skip_enrich:
        mask = _need_enrich(df)
        reste = int(mask.sum())
        print(f"== Enrichissement par lots : {reste} BE a traiter, lot={args.lot} ==", flush=True)
        if reste == 0:
            print("  tout est deja enrichi, on passe au scoring/export.", flush=True)
        else:
            # on prend le prochain lot
            idx_lot = df[mask].head(args.lot).index
            sub = df.loc[idx_lot].copy()
            print(f"  lot courant : {len(sub)} BE", flush=True)

            sub = enrich_sirene(session, sub)
            if time.time() - t0 < args.max_minutes * 60:
                sub = enrich_bodacc(session, sub)
            else:
                print("  (temps limite atteint apres SIRENE, BODACC au prochain run)", flush=True)
            try:
                sub = enrich_emails_sites(session, sub)
            except Exception as e:
                print(f"  emails interrompus ({e})", flush=True)

            sub["enrichi"] = "1"
            for c in ("etat_sirene", "tranche_effectif", "dirigeant",
                      "procedure_collective", "email", "email_source", "enrichi"):
                df.loc[idx_lot, c] = sub[c]

            df.to_csv(cache, index=False, quoting=csv.QUOTE_MINIMAL)
            done = int((df["enrichi"].astype(str) == "1").sum())
            print(f"  [sauvegarde] {done}/{len(df)} BE enrichis au total", flush=True)
            if done < len(df):
                print(f"  >>> Il reste {len(df)-done} BE. RELANCE le workflow pour continuer.", flush=True)

    # 2) Scoring + export (sur tout ce qu'on a, enrichi ou non)
    sc = df.apply(score_row, axis=1, result_type="expand")
    df["score"], df["score_detail"], df["effectif_num"] = sc[0], sc[1], sc[2]
    df["tier"] = df.apply(tier_row, axis=1)
    df["rk"] = df["tier"].map({t: i for i, t in enumerate(TIER_ORDER)})
    df = df.sort_values(["rk", "score"], ascending=[True, False]).drop(columns="rk")

    df.to_csv(cache, index=False, quoting=csv.QUOTE_MINIMAL)
    df[df["departement"].isin(DEPS_AURA)].to_csv(out / "base_be_max_aura.csv", index=False)
    df[df["departement"] == "69"].to_csv(out / "base_be_max_69.csv", index=False)
    try:
        build_excel(df, out / "base_be_tertiaire.xlsx")
    except Exception as e:
        print(f"  Excel non genere ({e}), les CSV sont disponibles.", flush=True)

    enrichi = int((df["enrichi"].astype(str) == "1").sum())
    summary = {
        "total": len(df),
        "enrichi": enrichi,
        "reste_a_enrichir": len(df) - enrichi,
        "avec_email": int(df["email"].astype(str).str.contains("@").sum()),
        "avec_tel": int((df["telephone"].astype(str).str.len() > 5).sum()),
        "aura": int(df["departement"].isin(DEPS_AURA).sum()),
        "dep_69": int((df["departement"] == "69").sum()),
        "tiers": df["tier"].value_counts().to_dict(),
    }
    (out / "run_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print("\n== RESULTATS ==")
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    if summary["reste_a_enrichir"] > 0:
        print(f"\n>>> {summary['reste_a_enrichir']} BE restants : relance le workflow pour continuer l'enrichissement.")
    else:
        print("\nEnrichissement COMPLET. Base nationale prete.")


if __name__ == "__main__":
    main()
